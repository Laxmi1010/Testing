import openpyxl

def get_data_from_excel(file_path, sheet_name):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]
    data = []

    for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip header row
        data.append(row)

    workbook.close()
    return data


#file_path → The path of the Excel file 
#sheet_name → The name of the worksheet inside the Excel file
#The function opens the Excel file using openpyxl.load_workbook().
#iter_rows() → Iterates over rows in the sheet.
#min_row=2 → Skips the header row (1st row).
#values_only=True → Returns cell values instead of Cell objects.
#row → A tuple representing data from a row.
#data.append(row) → Stores each row as a tuple inside the data list